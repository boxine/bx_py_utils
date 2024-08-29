import io
from unittest import TestCase

import openpyxl

from bx_py_utils.test_utils.snapshot import assert_text_snapshot
from bx_py_utils.test_utils.xlsx import FreezeXlsxTimes, generate_xlsx_md_snapshot, xlsx2dict
from bx_py_utils.test_utils.zip_file_utils import zip_info_markdown


class XlsxUtilsTestCase(TestCase):
    maxDiff = None

    @FreezeXlsxTimes()
    def test_xlsx2dict_simple(self):
        buffer = io.BytesIO()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet 1'
        ws.append(['Header 1', 'Header 2'])
        ws.append(['A1', 'B1'])
        ws.append(['A2', 'B2'])
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = buffer.getvalue()

        result = xlsx2dict(xlsx_data)
        self.assertEqual(
            result,
            {
                'Sheet 1': [
                    {'Header 1': 'A1', 'Header 2': 'B1'},
                    {'Header 1': 'A2', 'Header 2': 'B2'},
                ]
            },
        )

        # Demonstrates how to snapshot the XLSX zip file content:
        assert_text_snapshot(got=zip_info_markdown(xlsx_data), extension='.md')

    @FreezeXlsxTimes('2024-12-24T00:00:00+00:00')
    def test_xlsx2dict_complex(self):
        buffer = io.BytesIO()

        wb = openpyxl.Workbook()
        for sheet_name in ('New Sheet 1', 'New Sheet 2'):
            ws = wb.create_sheet(sheet_name)
            ws.title = sheet_name
            ws.append(['Header 1', 'Header 2'])
            ws.append(['A1', 'B1'])
            ws.append(['A2', 'B2'])
            ws.append(['A2', 'B2', 'C2'])  # Extra column
            ws.append(['A3', 'B3', 'C3', 'D4'])  # Two extra columns
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = buffer.getvalue()

        result = xlsx2dict(xlsx_data)
        self.assertEqual(
            result,
            {
                'New Sheet 1': [
                    {'Header 1': 'A1', 'Header 2': 'B1', 'UnnamedColumn3': None, 'UnnamedColumn4': None},
                    {'Header 1': 'A2', 'Header 2': 'B2', 'UnnamedColumn3': None, 'UnnamedColumn4': None},
                    {'Header 1': 'A2', 'Header 2': 'B2', 'UnnamedColumn3': 'C2', 'UnnamedColumn4': None},
                    {'Header 1': 'A3', 'Header 2': 'B3', 'UnnamedColumn3': 'C3', 'UnnamedColumn4': 'D4'},
                ],
                'New Sheet 2': [
                    {'Header 1': 'A1', 'Header 2': 'B1', 'UnnamedColumn3': None, 'UnnamedColumn4': None},
                    {'Header 1': 'A2', 'Header 2': 'B2', 'UnnamedColumn3': None, 'UnnamedColumn4': None},
                    {'Header 1': 'A2', 'Header 2': 'B2', 'UnnamedColumn3': 'C2', 'UnnamedColumn4': None},
                    {'Header 1': 'A3', 'Header 2': 'B3', 'UnnamedColumn3': 'C3', 'UnnamedColumn4': 'D4'},
                ],
                'Sheet': [],
            },
        )

        # Demonstrates how to snapshot the XLSX zip file content:
        assert_text_snapshot(got=generate_xlsx_md_snapshot(xlsx_data), extension='.md')
